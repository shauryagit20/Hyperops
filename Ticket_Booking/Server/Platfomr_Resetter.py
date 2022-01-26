import os
import openpyxl
import json
def reset():
    root   =  os.getcwd()
    wb = openpyxl.load_workbook("Schedule (2).xlsx")
    ws =  wb.active
    out_dict = {}
    for i in range(27,185):

        starting_values = ws[f"F{str(i)}"].value
        ending_values =  ws[f"G{str(i)}"].value
        print(f"The starting value is {starting_values}")
        key = f"{starting_values}-{ending_values}"
        print(key)
        out_dict[key] = {"Occupancy": 0, "Arrangement":{"Platform 1": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}, "Platform 2": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}, "Platform 3": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}, "Platform 4": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}, "Platform 5": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}, "Platform 6": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}, "Platform 8": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}, "Platform 9": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}, "Platform 10": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}, "Platform 11": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}, "Platform 12": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}, "Platform 13": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0}}}
    print(type(out_dict))

    with open("Timings.json","w") as f:
        json.dump(out_dict,f)
reset()